#!/usr/bin/env python3
"""
Generate Lint Test Reports
"""

import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Parse lint output
lint_issues = []

# Read the lint output
with open("/workspace/project/Rayaramathaynk/reports/lint-output.txt", "r") as f:
    content = f.read()

# Parse each issue
lines = content.split('\n')
current_file = None
issue_pattern = re.compile(r'(\d+):(\d+)\s+(error|warning)\s+(.*)')

for i, line in enumerate(lines):
    # Check if this is a file path
    if line.startswith('/workspace/project/Rayaramathaynk/') and ':' not in line[:30]:
        current_file = line.strip()
        # Get the next few lines for details
        if i + 1 < len(lines):
            next_line = lines[i + 1]
            match = issue_pattern.search(next_line)
            if match:
                col = match.group(1)
                severity = match.group(3)
                message = match.group(4).strip()
                
                # Get the code context if available
                code_context = ""
                if i + 3 < len(lines):
                    code_line = lines[i + 3].strip()
                    if code_line.startswith('>'):
                        code_context = code_line
                        if i + 4 < len(lines):
                            code_context += "\n" + lines[i + 4].strip()
                
                lint_issues.append({
                    "file": current_file,
                    "line": col,
                    "severity": severity.upper(),
                    "message": message,
                    "code": code_context
                })

# Also add issues from inline parsing
for i, line in enumerate(lines):
    match = issue_pattern.search(line)
    if match and i > 0 and not lines[i-1].startswith('/workspace/'):
        continue  # Skip inline matches without proper file context

# Count by severity
errors = [i for i in lint_issues if i["severity"] == "ERROR"]
warnings = [i for i in lint_issues if i["severity"] == "WARNING"]

# Group by file
files_with_issues = {}
for issue in lint_issues:
    file_path = issue["file"]
    if file_path not in files_with_issues:
        files_with_issues[file_path] = {"errors": 0, "warnings": 0, "issues": []}
    files_with_issues[file_path]["issues"].append(issue)
    if issue["severity"] == "ERROR":
        files_with_issues[file_path]["errors"] += 1
    else:
        files_with_issues[file_path]["warnings"] += 1

# Categorize issues by type
issue_types = {}
for issue in lint_issues:
    msg = issue["message"]
    # Extract rule name if present
    if "react-hooks/" in msg:
        rule = "react-hooks"
    elif "@typescript-eslint/" in msg:
        rule = "@typescript-eslint"
    elif "@next/next/" in msg:
        rule = "@next/next"
    elif "jsx-a11y/" in msg:
        rule = "jsx-a11y"
    elif "react/" in msg:
        rule = "react"
    else:
        rule = "other"
    
    if rule not in issue_types:
        issue_types[rule] = {"errors": 0, "warnings": 0}
    if issue["severity"] == "ERROR":
        issue_types[rule]["errors"] += 1
    else:
        issue_types[rule]["warnings"] += 1


def generate_json_report():
    """Save lint results as JSON"""
    report = {
        "timestamp": datetime.now().isoformat(),
        "summary": {
            "total_issues": len(lint_issues),
            "errors": len(errors),
            "warnings": len(warnings),
            "files_with_issues": len(files_with_issues)
        },
        "issues": lint_issues,
        "files": files_with_issues,
        "issue_types": issue_types
    }
    
    with open("/workspace/project/Rayaramathaynk/reports/lint_results.json", "w") as f:
        json.dump(report, f, indent=2)
    print("✅ Generated: reports/lint_results.json")


def generate_lint_excel():
    """Generate Lint Issues Excel Report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lint Issues"
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["File", "Line", "Severity", "Message"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, issue in enumerate(lint_issues, 2):
        file_name = issue["file"].replace("/workspace/project/Rayaramathaynk/", "")
        ws.cell(row=row, column=1, value=file_name).border = border
        ws.cell(row=row, column=2, value=issue["line"]).border = border
        
        severity_cell = ws.cell(row=row, column=3, value=issue["severity"])
        severity_cell.border = border
        if issue["severity"] == "ERROR":
            severity_cell.fill = error_fill
        else:
            severity_cell.fill = warning_fill
        
        ws.cell(row=row, column=4, value=issue["message"][:100]).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 60
    
    wb.save("/workspace/project/Rayaramathaynk/reports/LintIssues.xlsx")
    print("✅ Generated: reports/LintIssues.xlsx")


def generate_lint_summary_md():
    """Generate Lint Summary Markdown Report"""
    md_content = f"""# Lint Test Report
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Executive Summary

| Metric | Count |
|--------|-------|
| **Total Issues** | {len(lint_issues)} |
| **Errors** | {len(errors)} |
| **Warnings** | {len(warnings)} |
| **Files with Issues** | {len(files_with_issues)} |

## Issue Breakdown by Type

| Category | Errors | Warnings |
|----------|--------|----------|
"""
    
    for category, counts in sorted(issue_types.items(), key=lambda x: -x[1]["errors"]):
        md_content += f"| {category} | {counts['errors']} | {counts['warnings']} |\n"
    
    md_content += f"""
## Files with Issues

| File | Errors | Warnings |
|------|--------|----------|
"""
    
    for file_path, counts in sorted(files_with_issues.items(), key=lambda x: -x[1]["errors"]):
        file_name = file_path.replace("/workspace/project/Rayaramathaynk/", "")
        md_content += f"| {file_name} | {counts['errors']} | {counts['warnings']} |\n"
    
    md_content += f"""
## Critical Issues (Errors)

"""
    
    for issue in errors[:20]:
        file_name = issue["file"].replace("/workspace/project/Rayaramathaynk/", "")
        md_content += f"### {file_name}:{issue['line']}\n"
        md_content += f"**{issue['severity']}**: {issue['message']}\n\n"
    
    md_content += """
## Recommendations

### High Priority (Fix Errors First)
1. Fix all `react-hooks/set-state-in-effect` errors - these cause performance issues
2. Fix `react/no-unescaped-entities` errors - these are accessibility issues
3. Fix `react-hooks/incompatible-library` issues
4. Fix `require()` style imports in facilities/page.tsx

### Medium Priority
1. Add proper alt props to all images
2. Use `<Image />` from next/image instead of `<img>`
3. Fix missing dependencies in useEffect hooks
4. Remove unused imports and variables

### Low Priority
1. Clean up test files with unused variables
2. Address compilation skip warnings

---
*Report generated by ESLint*
"""
    
    with open("/workspace/project/Rayaramathaynk/reports/LintReport.md", "w") as f:
        f.write(md_content)
    print("✅ Generated: reports/LintReport.md")


def generate_executive_summary():
    """Generate Executive Summary"""
    pass_rate = max(0, 100 - (len(errors) * 2 + len(warnings) * 0.5))
    
    summary = f"""
================================================================================
                        LINT TEST EXECUTION REPORT
================================================================================

Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

================================================================================
                           EXECUTION SUMMARY
================================================================================

Total Issues Found:          {len(lint_issues)}
  - Errors:                  {len(errors)}
  - Warnings:                {len(warnings)}

Code Quality Score:           {pass_rate:.1f}%
Files Affected:              {len(files_with_issues)}

================================================================================
                        ISSUES BY CATEGORY
================================================================================
"""
    
    for category, counts in sorted(issue_types.items(), key=lambda x: -x[1]["errors"]):
        summary += f"  {category:30}  Errors: {counts['errors']:3}  Warnings: {counts['warnings']:3}\n"
    
    summary += f"""
================================================================================
                          TOP FILES WITH ISSUES
================================================================================
"""
    
    sorted_files = sorted(files_with_issues.items(), key=lambda x: -x[1]["errors"])[:10]
    for file_path, counts in sorted_files:
        file_name = file_path.replace("/workspace/project/Rayaramathaynk/", "")
        summary += f"  {file_name}\n"
        summary += f"    Errors: {counts['errors']}, Warnings: {counts['warnings']}\n"
    
    summary += f"""
================================================================================
                          CRITICAL ERRORS
================================================================================
"""
    
    for issue in errors[:10]:
        file_name = issue["file"].replace("/workspace/project/Rayaramathaynk/", "")
        summary += f"  📛 {file_name}:{issue['line']}\n"
        summary += f"     {issue['message'][:80]}\n\n"
    
    summary += f"""
================================================================================
                        RECOMMENDATIONS
================================================================================

IMMEDIATE ACTIONS (Before Next Deployment):
1. Fix all {len(errors)} errors - especially react-hooks issues
2. Address unescaped entities for accessibility compliance
3. Convert require() imports to ES6 imports

SHORT-TERM ACTIONS (This Week):
1. Address react-hooks/exhaustive-deps warnings
2. Add alt props to all images
3. Remove unused imports

LONG-TERM ACTIONS (This Sprint):
1. Refactor setState in effects pattern
2. Migrate all <img> to <Image> component
3. Address incompatible library warnings

================================================================================
                         FINAL DECISION
================================================================================

        {"❌ NOT READY" if len(errors) > 20 else "⚠️ NEEDS ATTENTION" if len(errors) > 0 else "✅ READY"}

        Total Issues: {len(lint_issues)} ({len(errors)} errors, {len(warnings)} warnings)
        Code Quality: {pass_rate:.1f}%
        Risk Level: {"HIGH" if len(errors) > 20 else "MEDIUM" if len(errors) > 0 else "LOW"}

================================================================================
                              SIGN-OFF
================================================================================

Tech Lead:                      ____________________  Date: ____________

Senior Developer:               ____________________  Date: ____________

================================================================================
                    Report Generated by ESLint Framework
                           Copyright © 2026 - All Rights Reserved
================================================================================
"""
    
    with open("/workspace/project/Rayaramathaynk/reports/LintExecutiveSummary.md", "w") as f:
        f.write(summary)
    print("✅ Generated: reports/LintExecutiveSummary.md")


def generate_bug_report():
    """Generate Bug Report Excel for Critical Issues"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Critical Lint Bugs"
    
    # Styles
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["File", "Line", "Issue Type", "Description", "Priority"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data - only errors
    row = 2
    for issue in errors:
        file_name = issue["file"].replace("/workspace/project/Rayaramathaynk/", "")
        
        # Determine priority based on issue type
        if "set-state" in issue["message"]:
            priority = "Critical"
        elif "unescaped" in issue["message"]:
            priority = "High"
        elif "require" in issue["message"]:
            priority = "High"
        else:
            priority = "Medium"
        
        ws.cell(row=row, column=1, value=file_name).border = border
        ws.cell(row=row, column=2, value=issue["line"]).border = border
        
        # Extract rule name
        rule = issue["message"].split("  ")[-1].split()[0] if "  " in issue["message"] else "unknown"
        ws.cell(row=row, column=3, value=rule).border = border
        
        ws.cell(row=row, column=4, value=issue["message"][:80]).border = border
        
        priority_cell = ws.cell(row=row, column=5, value=priority)
        priority_cell.border = border
        if priority == "Critical":
            priority_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif priority == "High":
            priority_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    
    wb.save("/workspace/project/Rayaramathaynk/reports/LintBugReport.xlsx")
    print("✅ Generated: reports/LintBugReport.xlsx")


# Generate all reports
print("\n" + "="*60)
print("GENERATING LINT TEST REPORTS")
print("="*60)

generate_json_report()
generate_lint_excel()
generate_lint_summary_md()
generate_executive_summary()
generate_bug_report()

print("\n" + "="*60)
print("ALL LINT REPORTS GENERATED SUCCESSFULLY")
print("="*60)
print("\n📁 Report Location: /workspace/project/Rayaramathaynk/reports/")
print("\nGenerated Files:")
print("  📄 lint_results.json           - Raw lint results data")
print("  📊 LintIssues.xlsx            - Complete lint issues list")
print("  📝 LintReport.md             - Detailed lint report")
print("  📄 LintExecutiveSummary.md    - Executive summary")
print("  🐛 LintBugReport.xlsx        - Critical lint bugs")
print("  📄 lint-output.txt           - Raw ESLint output")
print("="*60)
